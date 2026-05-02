from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='4A1D7A')
cat_font = Font(name='Arial', bold=True, color='1E1B4B', size=11)
cat_fill = PatternFill('solid', fgColor='EDE9FE')
dim_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
dim_fill = PatternFill('solid', fgColor='6B21A8')
normal_font = Font(name='Arial', size=10)
wrap_align = Alignment(wrap_text=True, vertical='center')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)

def setup_sheet(ws, title, headers):
    ws.title = title
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def add_dim(ws, row, dim_name, col_span=5):
    cell = ws.cell(row=row, column=1, value=dim_name)
    cell.font = dim_font
    cell.fill = dim_fill
    cell.alignment = center_align
    cell.border = thin_border
    for c in range(2, col_span+1):
        ws.cell(row=row, column=c).fill = dim_fill
        ws.cell(row=row, column=c).border = thin_border

def add_cat(ws, row, cat_name, col_span=5):
    cell = ws.cell(row=row, column=1, value=cat_name)
    cell.font = cat_font
    cell.fill = cat_fill
    cell.alignment = Alignment(vertical='center')
    cell.border = thin_border
    for c in range(2, col_span+1):
        ws.cell(row=row, column=c).fill = cat_fill
        ws.cell(row=row, column=c).border = thin_border

def add_row(ws, row, vals, start_col=1):
    for i, v in enumerate(vals, start_col):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = normal_font
        cell.alignment = wrap_align
        cell.border = thin_border

headers5 = ['序号', '大类', '细分类目', '说明', '代表品牌/案例']
headers3 = ['序号', '分类', '说明/代表']

# ===== Sheet 1: 按产品形态 =====
ws1 = wb.active
setup_sheet(ws1, '按产品形态细分', headers5)
ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 40
ws1.column_dimensions['E'].width = 42
ws1.freeze_panes = 'A2'

data_shape = [
    ('一、文具手帐', [
        (1, '文具手帐', '手帐本/笔记本', '活页本、定页本、日程本、bullet journal', '九口山、Kinbor、Moleskine'),
        (2, '文具手帐', '胶带/贴纸', '和纸胶带、手帐贴纸、装饰贴纸', 'MT胶带、仓井意匠、自制手帐胶带工作室'),
        (3, '文具手帐', '笔类', '钢笔、中性笔、彩色笔、马克笔', '百乐、LAMY（潮搭线）、晨光（文创线）'),
        (4, '文具手帐', '纸品类', '信纸、明信片、书签、卡片', '纸聘、独立插画师出品'),
        (5, '文具手帐', '收纳/配件', '笔袋、收纳袋、书立、笔筒', '国誉、MUJI（文创线）'),
        (6, '文具手帐', '印章', '橡皮章、火漆印章、定制印章', '独立手作印章工作室'),
    ]),
    ('二、潮玩/艺术玩具', [
        (7, '潮玩/艺术玩具', '盲盒', '盒装盲抽、系列盲盒、限定款', '泡泡玛特、52TOYS、寻找独角兽'),
        (8, '潮玩/艺术玩具', '手办/模型', '树脂手办、机甲模型、场景模型', '万代、Good Smile、独立手办师'),
        (9, '潮玩/艺术玩具', '艺术玩具', '设计师玩具、限量艺术玩具', 'KAWS、BE@RBRICK、Pop Mart（MEGA系列）'),
        (10, '潮玩/艺术玩具', '积木/拼装', '创意积木、拼装玩具', '乐高（创意线）、Keeppley'),
        (11, '潮玩/艺术玩具', '桌游/卡牌', '桌游、TCG卡牌、剧本杀道具', '狼人杀官方、独立桌游工作室'),
    ]),
    ('三、文创周边', [
        (12, '文创周边', '帆布袋', '原创设计帆布袋、联名款', '独立插画师工作室、博物馆文创店'),
        (13, '文创周边', '手机壳', '原创设计手机壳、IP联名壳', 'Casetify（中国设计线）、独立设计店铺'),
        (14, '文创周边', '马克杯/水杯', '创意杯型、原创图案杯', '故宫文创、独立陶瓷工作室'),
        (15, '文创周边', '钥匙扣/徽章', '金属徽章、珐琅徽章、亚克力挂件', '独立设计工作室、景区文创'),
        (16, '文创周边', '抱枕/靠垫', '原创图案抱枕、IP形象靠垫', '独立家居文创品牌'),
        (17, '文创周边', '雨伞', '创意伞面、联名伞', '独立设计品牌'),
        (18, '文创周边', '文创服饰', '文创T恤、卫衣、帽子', '优衣库UT（联名线）、独立插画师联名'),
    ]),
    ('四、IP衍生品', [
        (19, 'IP衍生品', '动漫IP周边', '国漫/日漫/美漫周边', '哔哩哔哩官方、腾讯动漫'),
        (20, 'IP衍生品', '游戏IP周边', '游戏手办、服装、道具', '米哈游（原神）、网易（阴阳师）'),
        (21, 'IP衍生品', '影视IP周边', '电影/电视剧联名周边', '万代南梦宫、迪士尼'),
        (22, 'IP衍生品', '文创IP', '博物馆/景区/城市IP', '故宫文创、敦煌文创、苏州博物馆'),
        (23, 'IP衍生品', '虚拟偶像周边', 'VTuber/虚拟人周边', 'A-SOUL、虚拟偶像工作室'),
        (24, 'IP衍生品', '品牌联名周边', '品牌跨界联名限量款', '瑞幸×茅台、喜茶×Fendi'),
    ]),
    ('五、艺术品/版画', [
        (25, '艺术品/版画', '限量版画', '签名版画、丝网版画、数码版画', '独立画廊、淘宝原画店'),
        (26, '艺术品/版画', '插画原画', '插画师原创作品出售', '站酷插画师、独立插画师'),
        (27, '艺术品/版画', '艺术微喷', '博物馆级复制画、装饰画', '佳能艺术微喷、独立艺术工作室'),
        (28, '艺术品/版画', '摄影作品', '限量签名摄影作品', '独立摄影师、图虫'),
        (29, '艺术品/版画', '雕塑/摆件', '小型艺术雕塑、桌面摆件', '独立艺术家、艺术商店'),
    ]),
    ('六、独立出版物', [
        (30, '独立出版物', '独立杂志', 'ZINE、独立刊物', '假杂志、Same Paper'),
        (31, '独立出版物', '摄影集', '独立摄影画册、影集', '独立摄影师出品'),
        (32, '独立出版物', '绘本/漫画', '独立绘本、原创漫画', '幽·灵、炭笔狗（微博漫画家）'),
        (33, '独立出版物', '艺术书', '艺术家书、实验出版物', '独立艺术书店出品'),
        (34, '独立出版物', '海报', '原创设计海报、电影海报收藏', '独立平面设计师'),
    ]),
    ('七、香薰/蜡烛', [
        (35, '香薰/蜡烛', '香氛蜡烛', '大豆蜡、蜂蜡、创意容器蜡烛', '观夏、Voluspa、野兽派'),
        (36, '香薰/蜡烛', '无火香薰', '藤条香薰、扩香石、香薰机', '观夏、野兽派'),
        (37, '香薰/蜡烛', '线香/盘香', '沉香、檀香、创意线香', '隐山溪、独立制香工作室'),
        (38, '香薰/蜡烛', '香薰礼盒', '节日礼盒、定制香薰套装', '各文创品牌节日线'),
    ]),
    ('八、创意家居装饰', [
        (39, '创意家居装饰', '创意摆件', '趣味摆件、桌面装饰', '造物集、吱音'),
        (40, '创意家居装饰', '创意灯饰', '设计感台灯、氛围灯', '东西设计、独立设计师品牌'),
        (41, '创意家居装饰', '创意挂画', '原创装饰画、磁吸画框', '独立插画师出品'),
        (42, '创意家居装饰', '创意花器', '设计花瓶、创意花盆', '野兽派、独立陶瓷工作室'),
        (43, '创意家居装饰', '创意钟表', '设计挂钟、趣味闹钟', '独立设计品牌'),
        (44, '创意家居装饰', '创意镜子', '造型镜、LED氛围镜', '独立家居品牌'),
    ]),
]

row = 2
for dim_name, items in data_shape:
    add_dim(ws1, row, dim_name)
    row += 1
    for item in items:
        add_row(ws1, row, item)
        row += 1

# ===== Sheet 2: 按文化主题 =====
ws2 = wb.create_sheet('按文化主题细分')
setup_sheet(ws2, '按文化主题细分', ['序号', '文化主题', '说明', '代表案例'])
ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 45
ws2.column_dimensions['D'].width = 40
ws2.freeze_panes = 'A2'

themes = [
    (1, '国潮/传统文化', '汉服、国风文创、非遗手作、传统文化元素再设计', '故宫文创、花西子、观夏'),
    (2, '博物馆文创', '博物馆联名/自主开发文创产品', '故宫、敦煌、苏州博物馆、三星堆'),
    (3, '城市文创', '城市IP、城市地标纪念品、地方文化产品', '南京盐水鸭文创、成都熊猫文创'),
    (4, '校园文创', '高校联名、毕业季纪念品、校园IP', '各高校文创店'),
    (5, '节庆文创', '春节/中秋/圣诞/七夕限定文创', '各品牌节日限定线'),
    (6, '二次元/ACG', '动漫、漫画、游戏周边衍生', '哔哩哔哩、米哈游'),
    (7, '复古怀旧', '复古玩具、怀旧包装、年代感设计', '老字号国潮线、回力'),
    (8, '科技感/赛博', '科技美学、未来感设计、赛博朋克风格', '独立科技文创工作室'),
    (9, '治愈系/可爱风', '萌系IP、治愈系产品、解压玩具', '线条小狗、Loopy、Jellycat'),
]
row = 2
for t in themes:
    add_row(ws2, row, t)
    row += 1

# ===== Sheet 3: 按消费场景 =====
ws3 = wb.create_sheet('按消费场景细分')
setup_sheet(ws3, '按消费场景细分', ['序号', '消费场景', '适合的产品类型', '目标人群'])
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 45
ws3.column_dimensions['D'].width = 30
ws3.freeze_panes = 'A2'

scenes = [
    (1, '办公桌面', '笔记本、笔、收纳、桌面摆件、马克杯', '职场白领、学生'),
    (2, '家居装饰', '挂画、花器、香薰、摆件、灯饰', '文艺青年、新婚家庭'),
    (3, '旅行出行', '帆布袋、旅行本、印章、冰箱贴', '旅行爱好者、游客'),
    (4, '送礼场景', '礼盒套装、限量版、定制款', '节日送礼、商务伴手礼'),
    (5, '收藏场景', '限量盲盒、手办、版画、签名原画', '深度玩家、收藏家'),
    (6, '日常陪伴', '钥匙扣、手机壳、徽章、贴纸', '泛年轻人群'),
    (7, '社交打卡', '高颜值产品、联名限量、网红款', '小红书/抖音用户'),
]
row = 2
for s in scenes:
    add_row(ws3, row, s)
    row += 1

# ===== Sheet 4: 按创作者类型（拓圈视角）=====
ws4 = wb.create_sheet('按创作者类型（拓圈视角）')
setup_sheet(ws4, '按创作者类型（拓圈视角）', ['序号', '创作者类型', '说明', '拓圈赋能方向'])
ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 40
ws4.column_dimensions['D'].width = 48
ws4.freeze_panes = 'A2'

creators = [
    (1, '独立插画师', '有原创IP/画风，做周边变现', '选品建议+供应链对接+OPC内容矩阵'),
    (2, '文创品牌主', '已有文创品牌，需要渠道和流量', '渠道铺设+品牌升级+OPC矩阵'),
    (3, '博物馆/景区文创团队', '有文化IP，缺设计+电商能力', '设计师对接+电商运营+OPC带货'),
    (4, '手作匠人', '非遗/手工/陶艺/木作等', '产品化包装+内容制作+商业对接'),
    (5, '独立设计师', '平面/产品/服装设计师做产品线', '产品孵化+供应链+商业化'),
    (6, 'IP持有方', '有原创IP形象/角色', 'IP商业化规划+周边开发+渠道对接'),
    (7, '高校文创团队', '师生文创项目', '商业化辅导+渠道资源+品牌打造'),
]
row = 2
for c in creators:
    add_row(ws4, row, c)
    row += 1

output = r'c:\Users\Administrator\WorkBuddy\Claw\文创产品类目深度细分.xlsx'
wb.save(output)
print(f'Excel已保存: {output}')
