from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
input_font = Font(color='0000FF')
calc_font = Font(color='000000')
note_fill = PatternFill('solid', fgColor='FFF2CC')
title_font = Font(bold=True, size=16)
section_font = Font(bold=True, size=12, color='4472C4')
sub_font = Font(bold=True, size=11, color='FFFFFF')
sub_fill = PatternFill('solid', fgColor='4472C4')
center = Alignment(horizontal='center')
right = Alignment(horizontal='right')

def sc(ws, r, c, v, font=None, fill=None, align=None, border=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if fmt: cell.number_format = fmt
    return cell

# ===== Sheet 1: 南京市场总览 =====
ws1 = wb.active
ws1.title = '南京市场总览'

sc(ws1, 1, 1, '拓圈·南京市场 目标与测算', title_font)
ws1.merge_cells('A1:F1')

sc(ws1, 3, 1, '一、价格体系（种子期）', section_font)
headers = ['角色', '费用项目', '金额（元）']
for i, h in enumerate(headers, 1):
    sc(ws1, 4, i, h, sub_font, sub_fill, center, thin_border)

price_data = [
    ['Wanderer漫游者', '免费', 0],
    ['Seeker探索者', '年费（种子期）', 99],
    ['Seeker探索者', '年费（正式期）', 199],
    ['Creator创造者', '年费（种子期）', 0],
    ['Creator创造者', '保证金（种子期）', 2000],
    ['Creator创造者', '年费（正式期）', 2999],
    ['Creator创造者', '保证金（正式期）', 5000],
    ['Worker工作者', '年费（种子期）', 0],
    ['Worker工作者', '年费（正式期）', 2999],
    ['Linker链接者', '保证金', 1000],
]
for i, (a, b, c) in enumerate(price_data, 5):
    sc(ws1, i, 1, a, calc_font, None, center, thin_border)
    sc(ws1, i, 2, b, calc_font, None, center, thin_border)
    sc(ws1, i, 3, c, input_font, None, center, thin_border, '#,##0')

sc(ws1, 17, 1, '二、南京第一年目标人数（按季度）', section_font)
ws1.merge_cells('A17:F17')
h2 = ['角色', '年目标', 'Q1', 'Q2', 'Q3', 'Q4']
for i, h in enumerate(h2, 1):
    sc(ws1, 18, i, h, sub_font, sub_fill, center, thin_border)

roles = [
    ['Wanderer', 40000, 4000, 10000, 14000, 12000],
    ['Seeker', 6500, 500, 1500, 2500, 2000],
    ['Creator', 300, 30, 80, 120, 70],
    ['Worker', 100, 10, 25, 40, 25],
    ['Linker', 30, 3, 8, 12, 7],
]
for i, row in enumerate(roles, 19):
    for j, v in enumerate(row, 1):
        fl = note_fill if j == 2 else None
        sc(ws1, i, j, v, input_font if isinstance(v, (int,float)) else calc_font, fl, center, thin_border, '#,##0' if j>=2 else None)

sc(ws1, 24, 1, '合计', Font(bold=True), None, center, thin_border)
for j in range(2, 7):
    cl = get_column_letter(j)
    sc(ws1, 24, j, f'=SUM({cl}19:{cl}23)', Font(bold=True), None, center, thin_border, '#,##0')

sc(ws1, 26, 1, '三、核心假设参数（蓝色=可调）', section_font)
ws1.merge_cells('A26:D26')
assumptions = [
    ['平台抽成比例', '8%', '创造者交易额的抽成'],
    ['Linker分润比例', '1.5%', '从平台抽成中支出'],
    ['创造者月均交易额-门店型', 5000, '咖啡店/茶叶店等'],
    ['创造者月均交易额-产品型', 8000, '实物/线上产品'],
    ['创造者月均交易额-活动型', 15000, '单场活动收入'],
    ['创造者月均交易额-内容型', 3000, '课程/咨询等'],
    ['创造者月均交易额-服务型', 6000, '设计/策划等'],
    ['活跃创造者比例', '60%', '有实际交易的创造者占比'],
    ['Seeker转化购买率', '5%', '探索者中产生购买行为的比例'],
    ['Seeker客单价', 200, '探索者平均消费金额'],
]
for i, (p, v, n) in enumerate(assumptions, 27):
    sc(ws1, i, 1, p, Font(bold=True), None, None, thin_border)
    sc(ws1, i, 2, v, input_font, note_fill, center, thin_border)
    sc(ws1, i, 3, n, Font(size=10, color='666666'), None, None, thin_border)

for c, w in {'A':28,'B':22,'C':16,'D':16,'E':16,'F':16}.items():
    ws1.column_dimensions[c].width = w

# ===== Sheet 2: 季度营收测算 =====
ws2 = wb.create_sheet('季度营收测算')
sc(ws2, 1, 1, '拓圈·南京市场 季度营收测算（第一年）', title_font)
ws2.merge_cells('A1:G1')

sc(ws2, 3, 1, '收入测算', section_font)
h3 = ['收入来源', '计算逻辑', 'Q1', 'Q2', 'Q3', 'Q4', '全年合计']
for i, h in enumerate(h3, 1):
    sc(ws2, 4, i, h, sub_font, sub_fill, center, thin_border)

# Creator按类型分配: 门店30% 产品30% 活动15% 内容15% 服务10%
# 活跃比例60%, Q1-Q4累计创造者
# 交易抽成 = 活跃数 × 月均交易额 × 8% × 3个月(每季度)
# Q1:累计30, Q2:累计110, Q3:累计230, Q4:累计300

rev = [
    ['Seeker年费收入', '累计Seeker × 99元', 49500, 198000, 445500, 643500],
    ['创造者保证金（流入）', '新增Creator × 2000', 60000, 160000, 240000, 140000],
    ['Linker保证金（流入）', '新增Linker × 1000', 3000, 8000, 12000, 7000],
    ['交易抽成-门店型(30%)', '活跃门店×5000×6.5%×3月', 17550, 64350, 134550, 87750],
    ['交易抽成-产品型(30%)', '活跃产品×8000×6.5%×3月', 28080, 102960, 215280, 140400],
    ['交易抽成-活动型(15%)', '活跃活动×15000×6.5%×3月', 13163, 48394, 100980, 65813],
    ['交易抽成-内容型(15%)', '活跃内容×3000×6.5%×3月', 2633, 9679, 20196, 13163],
    ['交易抽成-服务型(10%)', '活跃服务×6000×6.5%×3月', 3510, 12888, 26910, 17550],
    ['Seeker消费收入', '累计Seeker × 5% × 200', 5000, 15000, 25000, 20000],
]

for i, (name, logic, q1, q2, q3, q4) in enumerate(rev, 5):
    sc(ws2, i, 1, name, calc_font, None, None, thin_border)
    sc(ws2, i, 2, logic, Font(size=9, color='666666'), None, None, thin_border)
    for j, v in enumerate([q1, q2, q3, q4], 3):
        sc(ws2, i, j, v, calc_font, None, right, thin_border, '#,##0')
    sc(ws2, i, 7, f'=SUM(C{i}:F{i})', Font(bold=True), None, right, thin_border, '#,##0')

tr = 14  # total row
sc(ws2, tr, 1, '收入合计', Font(bold=True, size=11), PatternFill('solid', fgColor='D9E2F3'), None, thin_border)
for j in range(3, 8):
    cl = get_column_letter(j)
    sc(ws2, tr, j, f'=SUM({cl}5:{cl}13)', Font(bold=True, size=11), PatternFill('solid', fgColor='D9E2F3'), right, thin_border, '#,##0')

cs = 16
sc(ws2, cs, 1, '成本测算', section_font)
for i, h in enumerate(h3, 1):
    sc(ws2, cs+1, i, h, sub_font, sub_fill, center, thin_border)

costs = [
    ['小程序认证', '300元/年', 300, 0, 0, 0],
    ['服务器/云服务', 'LeanCloud/Supabase', 150, 200, 300, 300],
    ['域名+托管', 'GitHub Pages+域名', 100, 0, 0, 0],
    ['内容制作', '视频/设计/文案', 5000, 8000, 10000, 8000],
    ['活动运营', '拓圈无界活动', 8000, 15000, 20000, 15000],
    ['Linker分润支出', '交易抽成×1.5%/6.5%', None, None, None, None],
]

for i, (name, desc, q1, q2, q3, q4) in enumerate(costs, cs+2):
    sc(ws2, i, 1, name, calc_font, None, None, thin_border)
    sc(ws2, i, 2, desc, Font(size=9, color='666666'), None, None, thin_border)
    if name == 'Linker分润支出':
        for j in range(3, 7):
            cl = get_column_letter(j)
            sc(ws2, i, j, f'=({cl}9+{cl}10+{cl}11+{cl}12+{cl}13)*1.5/6.5', input_font, None, right, thin_border, '#,##0')
        sc(ws2, i, 7, f'=SUM(C{i}:F{i})', Font(bold=True), None, right, thin_border, '#,##0')
    else:
        for j, v in enumerate([q1, q2, q3, q4], 3):
            sc(ws2, i, j, v, input_font, None, right, thin_border, '#,##0')
        sc(ws2, i, 7, f'=SUM(C{i}:F{i})', Font(bold=True), None, right, thin_border, '#,##0')

ce = cs + 2 + len(costs) - 1  # 22
sc(ws2, ce+1, 1, '成本合计', Font(bold=True, size=11), PatternFill('solid', fgColor='FCE4D6'), None, thin_border)
for j in range(3, 8):
    cl = get_column_letter(j)
    sc(ws2, ce+1, j, f'=SUM({cl}{cs+2}:{cl}{ce})', Font(bold=True, size=11), PatternFill('solid', fgColor='FCE4D6'), right, thin_border, '#,##0')

pr = ce + 3
sc(ws2, pr, 1, '季度净利润（扣除Linker分润后）', Font(bold=True, size=12), PatternFill('solid', fgColor='C6EFCE'), None, thin_border)
for j in range(3, 8):
    cl = get_column_letter(j)
    sc(ws2, pr, j, f'={cl}{tr}-{cl}{ce+1}', Font(bold=True, size=12), PatternFill('solid', fgColor='C6EFCE'), right, thin_border, '#,##0')

for c, w in {'A':32,'B':30,'C':14,'D':14,'E':14,'F':14,'G':16}.items():
    ws2.column_dimensions[c].width = w

# ===== Sheet 3: 创造者投资回报 =====
ws3 = wb.create_sheet('创造者投资回报')
sc(ws3, 1, 1, '创造者投资回报测算（种子期·产品型基准）', title_font)
ws3.merge_cells('A1:F1')

sc(ws3, 3, 1, '一、投入', section_font)
for i, h in enumerate(['项目', '金额（元）', '说明'], 1):
    sc(ws3, 4, i, h, sub_font, sub_fill, center, thin_border)

inv = [
    ['保证金', 2000, '退出时可退'],
    ['年费（种子期）', 0, '种子期免费'],
    ['时间投入', '-', '发内容/参加活动'],
    ['总投入（资金）', 2000, '实际资金投入'],
]
for i, (a, b, c) in enumerate(inv, 5):
    fl = note_fill if a=='总投入（资金）' else None
    sc(ws3, i, 1, a, Font(bold=True) if fl else calc_font, None, None, thin_border)
    sc(ws3, i, 2, b, input_font if isinstance(b,(int,float)) else calc_font, fl, center, thin_border, '#,##0' if isinstance(b,(int,float)) else None)
    sc(ws3, i, 3, c, Font(size=10, color='666666'), None, None, thin_border)

sc(ws3, 10, 1, '二、各类型创造者月收益预估', section_font)
ws3.merge_cells('A10:F10')

rh = ['创造者类型', '月均交易额', '到手比例(扣Linker)', '月净收入', '年净收入', '年ROI']
for i, h in enumerate(rh, 1):
    sc(ws3, 11, i, h, sub_font, sub_fill, center, thin_border)

types = [
    ['门店型（咖啡/茶/手作）', 5000],
    ['产品型（实物/线上）', 8000],
    ['活动型（展览/市集）', 15000],
    ['内容型（课程/咨询）', 3000],
    ['服务型（设计/策划）', 6000],
]
for i, (name, m) in enumerate(types, 12):
    sc(ws3, i, 1, name, calc_font, None, None, thin_border)
    sc(ws3, i, 2, m, input_font, None, right, thin_border, '#,##0')
    sc(ws3, i, 3, 0.065, input_font, None, center, thin_border, '0.0%')
    sc(ws3, i, 4, f'=B{i}*C{i}', calc_font, None, right, thin_border, '#,##0')
    sc(ws3, i, 5, f'=D{i}*12', calc_font, None, right, thin_border, '#,##0')
    sc(ws3, i, 6, f'=E{i}/2000', calc_font, PatternFill('solid', fgColor='E2EFDA'), center, thin_border, '0.0%')

sc(ws3, 18, 1, '三、号段增值预估（以产品型为基准）', section_font)
ws3.merge_cells('A18:D18')

for i, h in enumerate(['项目', '保守', '中性', '乐观'], 1):
    sc(ws3, 19, i, h, sub_font, sub_fill, center, thin_border)

sc(ws3, 20, 1, '1年后号段二手交易价', calc_font, None, None, thin_border)
for j, v in enumerate([3000, 5000, 10000], 2):
    sc(ws3, 20, j, v, input_font, note_fill, right, thin_border, '#,##0')

sc(ws3, 21, 1, '号段增值收益', calc_font, None, None, thin_border)
for j in range(2, 5):
    cl = get_column_letter(j)
    sc(ws3, 21, j, f'={cl}20-2000', calc_font, None, right, thin_border, '#,##0')

sc(ws3, 22, 1, '年净收入（产品型）', calc_font, None, None, thin_border)
for j in range(2, 5):
    sc(ws3, 22, j, f'=E13', calc_font, None, right, thin_border, '#,##0')

sc(ws3, 23, 1, '总回报（收益+号段增值）', Font(bold=True), PatternFill('solid', fgColor='D9E2F3'), None, thin_border)
for j in range(2, 5):
    cl = get_column_letter(j)
    sc(ws3, 23, j, f'={cl}22+{cl}21', Font(bold=True), PatternFill('solid', fgColor='D9E2F3'), right, thin_border, '#,##0')

sc(ws3, 24, 1, '总投资回报率', Font(bold=True), PatternFill('solid', fgColor='C6EFCE'), None, thin_border)
for j in range(2, 5):
    cl = get_column_letter(j)
    sc(ws3, 24, j, f'={cl}23/2000', Font(bold=True), PatternFill('solid', fgColor='C6EFCE'), center, thin_border, '0.0%')

sc(ws3, 26, 1, '结论：即使最保守估计，创造者投入2000元，1年后总回报率也超过100%', Font(bold=True, size=11, color='006100'))

for c, w in {'A':32,'B':18,'C':18,'D':16,'E':16,'F':14}.items():
    ws3.column_dimensions[c].width = w

# ===== Sheet 4: Linker收益测算 =====
ws4 = wb.create_sheet('Linker收益测算')
sc(ws4, 1, 1, 'Linker链接者 收益测算', title_font)
ws4.merge_cells('A1:F1')

sc(ws4, 3, 1, '一、Linker投入', section_font)
for i, h in enumerate(['项目', '金额（元）'], 1):
    sc(ws4, 4, i, h, sub_font, sub_fill, center, thin_border)
for i, (a, b) in enumerate([['保证金', 1000], ['时间投入', '拉人/组织活动']], 5):
    sc(ws4, i, 1, a, calc_font, None, None, thin_border)
    sc(ws4, i, 2, b, input_font if isinstance(b,(int,float)) else calc_font, note_fill if isinstance(b,(int,float)) else None, center, thin_border, '#,##0' if isinstance(b,(int,float)) else None)

sc(ws4, 8, 1, '二、Linker分级收益测算', section_font)
ws4.merge_cells('A8:F8')
lh = ['级别', '累计介绍客户', '分润比例', '预估月分润', '预估年分润', '年ROI']
for i, h in enumerate(lh, 1):
    sc(ws4, 9, i, h, sub_font, sub_fill, center, thin_border)

ll = [
    ['初级', '1-10人', '1.0%', 300, 3600],
    ['中级', '11-30人', '1.5%', 900, 10800],
    ['高级', '31-50人', '2.0%', 1800, 21600],
    ['顶级(城市分成)', '51人+', '2.5%+城市分成', 5000, 60000],
]
for i, (a, b, c, d, e) in enumerate(ll, 10):
    sc(ws4, i, 1, a, calc_font, None, center, thin_border)
    sc(ws4, i, 2, b, calc_font, None, center, thin_border)
    sc(ws4, i, 3, c, input_font, None, center, thin_border)
    sc(ws4, i, 4, d, input_font, None, right, thin_border, '#,##0')
    sc(ws4, i, 5, e, input_font, None, right, thin_border, '#,##0')
    sc(ws4, i, 6, f'=E{i}/1000', calc_font, PatternFill('solid', fgColor='E2EFDA'), center, thin_border, '0.0%')

sc(ws4, 15, 1, '三、Linker升级路径', section_font)
ws4.merge_cells('A15:C15')
for i, h in enumerate(['任务类型', '经验值', '说明'], 1):
    sc(ws4, 16, i, h, sub_font, sub_fill, center, thin_border)

tasks = [
    ['介绍1个新Seeker/Creator', 100, '主要升级方式'],
    ['组织1场线下活动', 500, '活动签到人数越多经验越高'],
    ['发布优质内容', 50, '被推荐/转发额外加经验'],
    ['促成1笔交易撮合', 200, '成功撮合买卖双方'],
    ['邀请大佬入驻', 1000, '高权重创造者/品牌'],
]
for i, (a, b, c) in enumerate(tasks, 17):
    sc(ws4, i, 1, a, calc_font, None, None, thin_border)
    sc(ws4, i, 2, b, input_font, None, center, thin_border, '#,##0')
    sc(ws4, i, 3, c, Font(size=10, color='666666'), None, None, thin_border)

sc(ws4, 23, 1, '注：分润金额为预估值，取决于介绍客户的实际交易额', Font(size=9, color='999999'))

for c, w in {'A':28,'B':20,'C':16,'D':16,'E':16,'F':14}.items():
    ws4.column_dimensions[c].width = w

output = r'C:\Users\Administrator\WorkBuddy\Claw\拓圈·南京市场测算模型.xlsx'
wb.save(output)
print(f'Done: {output}')
